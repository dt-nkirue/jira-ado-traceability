"""
Jira-ADO Traceability Report Generator (Scheduled Version)
Designed to run as a scheduled task/robot
"""

import json
import os
import sys
import logging
import traceback
from datetime import datetime
from pathlib import Path
import pandas as pd
import requests
from requests.auth import HTTPBasicAuth
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from fuzzywuzzy import fuzz
from fuzzywuzzy import process


class JiraAdoTraceabilityRobot:
    """Main robot class for Jira-ADO traceability reporting"""

    def __init__(self, config_path=None):
        """Initialize the robot with configuration"""
        self.config_path = config_path or os.path.join(os.path.dirname(__file__), 'config.json')
        self.config = self.load_config()
        self.setup_logging()
        self.logger.info("="*80)
        self.logger.info("Jira-ADO Traceability Robot Starting")
        self.logger.info(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.logger.info("="*80)

    def load_config(self):
        """Load configuration from file and environment variables"""
        if not os.path.exists(self.config_path):
            raise FileNotFoundError(f"Configuration file not found: {self.config_path}")

        with open(self.config_path, 'r') as f:
            config = json.load(f)

        # Override with environment variables if present
        config['ado']['pat'] = os.environ.get('ADO_PAT', config['ado']['pat'])
        config['jira']['data_file'] = os.environ.get('JIRA_DATA_FILE', config['jira']['data_file'])

        # Override email credentials with environment variables if present
        if config.get('notifications', {}).get('enabled', False):
            email_config = config['notifications']['email']
            email_config['username'] = os.environ.get('SMTP_USERNAME', email_config.get('username', ''))
            email_config['password'] = os.environ.get('SMTP_PASSWORD', email_config.get('password', ''))

        return config

    def setup_logging(self):
        """Setup logging to file and console"""
        log_dir = Path(self.config['logging']['log_directory'])
        log_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        log_file = log_dir / f"jira_ado_traceability_{timestamp}.log"

        # Create logger
        self.logger = logging.getLogger('JiraAdoRobot')
        self.logger.setLevel(self.config['logging']['log_level'])

        # File handler
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(self.config['logging']['log_level'])

        # Console handler
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.INFO)

        # Formatter
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)

        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)

        self.logger.info(f"Logging initialized. Log file: {log_file}")

    def load_jira_data(self):
        """Load Jira data from JSON file"""
        try:
            jira_file = self.config['jira']['data_file']
            self.logger.info(f"Loading Jira data from: {jira_file}")

            if not os.path.exists(jira_file):
                raise FileNotFoundError(f"Jira data file not found: {jira_file}")

            with open(jira_file, 'r', encoding='utf-8') as f:
                jira_data = json.load(f)

            jira_issues = jira_data.get('issues', [])
            self.logger.info(f"Loaded {len(jira_issues)} Jira issues")
            return jira_issues

        except Exception as e:
            self.logger.error(f"Failed to load Jira data: {str(e)}")
            raise

    def parse_jira_issues(self, jira_issues):
        """Parse Jira issues into structured data"""
        self.logger.info("Parsing Jira issues...")
        parsed_data = []

        for issue in jira_issues:
            try:
                fields = issue.get('fields', {})

                # Basic fields
                jira_key = issue.get('key', '')
                jira_summary = fields.get('summary', '')

                # Status
                status = fields.get('status', {})
                jira_status = status.get('name', 'Unknown')
                jira_status_category = status.get('statusCategory', {}).get('name', 'Unknown')

                # Priority
                priority = fields.get('priority', {})
                jira_priority = priority.get('name', 'None') if priority else 'None'

                # Severity
                severity = fields.get('customfield_10042', {})
                jira_severity = severity.get('value', 'None') if severity else 'None'

                # Assignee
                assignee = fields.get('assignee', {})
                jira_assignee = assignee.get('displayName', 'Unassigned') if assignee else 'Unassigned'

                # Dates
                created = fields.get('created', '')
                jira_created = pd.to_datetime(created, utc=True).tz_localize(None) if created else None

                resolutiondate = fields.get('resolutiondate', None)
                jira_resolved = pd.to_datetime(resolutiondate, utc=True).tz_localize(None) if resolutiondate else None

                # ADO fields
                ado_id = fields.get('customfield_10109', '')
                ado_state_jira = fields.get('customfield_10110', '')

                parsed_data.append({
                    'Jira Key': jira_key,
                    'Jira Summary': jira_summary,
                    'Jira Status': jira_status,
                    'Jira Status Category': jira_status_category,
                    'Jira Priority': jira_priority,
                    'Jira Severity': jira_severity,
                    'Jira Assignee': jira_assignee,
                    'Jira Created': jira_created,
                    'Jira Resolved': jira_resolved,
                    'ADO ID': ado_id if ado_id else 'Not Linked',
                    'ADO State (Jira)': ado_state_jira if ado_state_jira else 'N/A'
                })

            except Exception as e:
                self.logger.warning(f"Error parsing Jira issue {issue.get('key', 'unknown')}: {str(e)}")
                continue

        self.logger.info(f"Successfully parsed {len(parsed_data)} Jira issues")
        return pd.DataFrame(parsed_data)

    def fetch_ado_work_item(self, ado_id, auth):
        """Fetch a single ADO work item with retry logic"""
        max_retries = 3
        retry_count = 0

        while retry_count < max_retries:
            try:
                ado_config = self.config['ado']
                api_base = f"{ado_config['server']}/{ado_config['collection']}/{ado_config['project']}/_apis"
                url = f"{api_base}/wit/workitems/{ado_id}?api-version=5.0"

                response = requests.get(url, auth=auth, timeout=10)

                if response.status_code == 200:
                    work_item = response.json()
                    fields = work_item.get('fields', {})

                    return {
                        'ADO Title': fields.get('System.Title', ''),
                        'ADO State': fields.get('System.State', ''),
                        'ADO Assigned To': fields.get('System.AssignedTo', {}).get('displayName', 'Unassigned')
                            if isinstance(fields.get('System.AssignedTo'), dict)
                            else str(fields.get('System.AssignedTo', 'Unassigned')),
                        'ADO Work Item Type': fields.get('System.WorkItemType', ''),
                        'ADO Priority': str(fields.get('Microsoft.VSTS.Common.Priority', '')),
                        'ADO Severity': str(fields.get('Microsoft.VSTS.Common.Severity', '')),
                        'ADO Created Date': fields.get('System.CreatedDate', ''),
                        'ADO Closed Date': fields.get('Microsoft.VSTS.Common.ClosedDate', ''),
                        'ADO Resolved Date': fields.get('Microsoft.VSTS.Common.ResolvedDate', ''),
                        'ADO Area Path': fields.get('System.AreaPath', ''),
                        'ADO Iteration Path': fields.get('System.IterationPath', '')
                    }
                else:
                    self.logger.warning(f"Failed to fetch ADO-{ado_id}: HTTP {response.status_code}")
                    return self.create_error_ado_item(f"HTTP {response.status_code}")

            except requests.Timeout:
                retry_count += 1
                self.logger.warning(f"Timeout fetching ADO-{ado_id}, retry {retry_count}/{max_retries}")
                if retry_count >= max_retries:
                    return self.create_error_ado_item("Timeout after retries")

            except Exception as e:
                self.logger.error(f"Error fetching ADO-{ado_id}: {str(e)}")
                return self.create_error_ado_item(str(e))

        return self.create_error_ado_item("Max retries exceeded")

    def create_error_ado_item(self, error_msg):
        """Create error placeholder for failed ADO fetch"""
        return {
            'ADO Title': f'ERROR: {error_msg}',
            'ADO State': 'Error',
            'ADO Assigned To': '',
            'ADO Work Item Type': '',
            'ADO Priority': '',
            'ADO Severity': '',
            'ADO Created Date': '',
            'ADO Closed Date': '',
            'ADO Resolved Date': '',
            'ADO Area Path': '',
            'ADO Iteration Path': ''
        }

    def fetch_ado_work_items(self, df):
        """Fetch all linked ADO work items"""
        self.logger.info("Fetching Azure DevOps work items...")

        df_with_ado = df[df['ADO ID'] != 'Not Linked'].copy()
        ado_ids = df_with_ado['ADO ID'].unique()

        self.logger.info(f"Found {len(ado_ids)} unique ADO work items to fetch")

        ado_work_items = {}
        auth = HTTPBasicAuth('', self.config['ado']['pat'])

        for idx, ado_id in enumerate(ado_ids, 1):
            if not ado_id or ado_id == 'Not Linked':
                continue

            self.logger.info(f"Fetching ADO work item {ado_id} ({idx}/{len(ado_ids)})...")
            ado_work_items[str(ado_id)] = self.fetch_ado_work_item(ado_id, auth)

        self.logger.info(f"Successfully fetched {len(ado_work_items)} ADO work items")
        return ado_work_items

    def perform_fuzzy_matching(self, df):
        """Perform fuzzy matching for unlinked Jira items"""
        if not self.config['fuzzy_matching']['enabled']:
            self.logger.info("Fuzzy matching is disabled in configuration")
            return pd.DataFrame()

        self.logger.info("Performing fuzzy matching for unlinked Jira items...")

        try:
            ado_config = self.config['ado']
            api_base = f"{ado_config['server']}/{ado_config['collection']}/{ado_config['project']}/_apis"
            auth = HTTPBasicAuth('', ado_config['pat'])

            days_lookback = self.config['fuzzy_matching']['days_lookback']
            wiql_query = {
                'query': f"SELECT [System.Id], [System.Title] FROM WorkItems WHERE [System.TeamProject] = '{ado_config['project']}' AND [System.CreatedDate] >= @Today - {days_lookback} ORDER BY [System.CreatedDate] DESC"
            }

            wiql_url = f"{api_base}/wit/wiql?api-version=5.0"
            wiql_response = requests.post(wiql_url, auth=auth, json=wiql_query, timeout=30)

            all_ado_work_items = {}

            if wiql_response.status_code == 200:
                wiql_results = wiql_response.json()
                work_item_ids = [str(item['id']) for item in wiql_results.get('workItems', [])]
                self.logger.info(f"Found {len(work_item_ids)} ADO work items for potential matching")

                max_items = self.config['fuzzy_matching']['max_ado_items']
                for ado_id in work_item_ids[:max_items]:
                    try:
                        url = f"{api_base}/wit/workitems/{ado_id}?api-version=5.0"
                        response = requests.get(url, auth=auth, timeout=10)

                        if response.status_code == 200:
                            work_item = response.json()
                            fields = work_item.get('fields', {})
                            all_ado_work_items[ado_id] = {
                                'ADO Title': fields.get('System.Title', ''),
                                'ADO State': fields.get('System.State', ''),
                                'ADO Work Item Type': fields.get('System.WorkItemType', '')
                            }
                    except Exception as e:
                        continue

            # Perform fuzzy matching
            fuzzy_matches = []
            unlinked_items = df[df['ADO ID'] == 'Not Linked']
            min_score = self.config['fuzzy_matching']['min_score']

            if len(all_ado_work_items) > 0 and len(unlinked_items) > 0:
                self.logger.info(f"Analyzing {len(unlinked_items)} unlinked Jira items...")

                ado_titles = {ado_id: info['ADO Title'] for ado_id, info in all_ado_work_items.items()}

                for idx, jira_row in unlinked_items.iterrows():
                    jira_summary = jira_row['Jira Summary']
                    matches = process.extract(jira_summary, ado_titles, scorer=fuzz.token_sort_ratio, limit=5)

                    for match in matches:
                        ado_title, score, ado_id_key = match[0], match[1], None

                        for aid, title in ado_titles.items():
                            if title == ado_title:
                                ado_id_key = aid
                                break

                        if score >= min_score and ado_id_key:
                            confidence = 'Very High' if score >= 90 else 'High' if score >= 80 else 'Medium'

                            fuzzy_matches.append({
                                'Jira Key': jira_row['Jira Key'],
                                'Jira Summary': jira_summary,
                                'Jira Status': jira_row['Jira Status'],
                                'Potential ADO ID': ado_id_key,
                                'ADO Title': ado_title,
                                'ADO State': all_ado_work_items[ado_id_key]['ADO State'],
                                'ADO Work Item Type': all_ado_work_items[ado_id_key]['ADO Work Item Type'],
                                'Match Score': score,
                                'Confidence': confidence
                            })

            self.logger.info(f"Found {len(fuzzy_matches)} potential matches")
            return pd.DataFrame(fuzzy_matches)

        except Exception as e:
            self.logger.error(f"Error during fuzzy matching: {str(e)}")
            return pd.DataFrame()

    def merge_ado_data(self, df, ado_work_items):
        """Merge ADO data with Jira DataFrame"""
        self.logger.info("Merging Jira and ADO data...")

        ado_columns = ['ADO Title', 'ADO State', 'ADO Assigned To', 'ADO Work Item Type',
                      'ADO Priority', 'ADO Severity', 'ADO Created Date', 'ADO Closed Date',
                      'ADO Resolved Date', 'ADO Area Path', 'ADO Iteration Path']

        for col in ado_columns:
            df[col] = ''

        for idx, row in df.iterrows():
            ado_id = str(row['ADO ID'])
            if ado_id in ado_work_items:
                for col in ado_columns:
                    df.at[idx, col] = ado_work_items[ado_id].get(col, '')

        return df

    @staticmethod
    def compare_status(jira_status_cat, ado_state):
        """Compare status between Jira and ADO"""
        if not ado_state or ado_state == '':
            return 'No ADO Link'

        jira_done = jira_status_cat.lower() == 'done'
        ado_closed = ado_state.lower() in ['closed', 'resolved', 'done', 'removed']

        if jira_done and ado_closed:
            return '[OK] Both Closed'
        elif not jira_done and not ado_closed:
            return '[OK] Both Open'
        elif jira_done and not ado_closed:
            return '[WARN] Jira Closed, ADO Open'
        else:
            return '[WARN] ADO Closed, Jira Open'

    @staticmethod
    def compare_severity(jira_sev, ado_sev):
        """Compare severity between systems"""
        if not ado_sev or ado_sev == '':
            return 'N/A'

        jira_num = ''.join(filter(str.isdigit, str(jira_sev)))
        ado_num = str(ado_sev).strip()

        if jira_num == ado_num:
            return '[OK] Match'
        else:
            return f'[WARN] Mismatch (J:{jira_sev} vs A:{ado_sev})'

    def perform_comparison(self, df):
        """Perform comparison analysis"""
        self.logger.info("Performing comparison analysis...")

        df['Status Comparison'] = df.apply(
            lambda row: self.compare_status(row['Jira Status Category'], row['ADO State']),
            axis=1
        )
        df['Severity Comparison'] = df.apply(
            lambda row: self.compare_severity(row['Jira Severity'], row['ADO Severity']),
            axis=1
        )
        df['Assignee Match'] = df.apply(
            lambda row: '[OK] Match' if row['Jira Assignee'].lower() == row['ADO Assigned To'].lower()
            else '[WARN] Different',
            axis=1
        )

        return df

    def generate_summary_stats(self, df):
        """Generate summary statistics"""
        self.logger.info("Generating summary statistics...")

        total_issues = len(df)
        linked_issues = len(df[df['ADO ID'] != 'Not Linked'])
        unlinked_issues = total_issues - linked_issues

        status_mismatches = len(df[df['Status Comparison'].str.contains('[WARN]', na=False)])
        severity_mismatches = len(df[df['Severity Comparison'].str.contains('[WARN]', na=False)])
        assignee_mismatches = len(df[df['Assignee Match'].str.contains('[WARN]', na=False)])

        both_closed = len(df[df['Status Comparison'] == '[OK] Both Closed'])
        both_open = len(df[df['Status Comparison'] == '[OK] Both Open'])

        summary_stats = {
            'Metric': [
                'Total Jira Issues',
                'Linked to ADO',
                'Not Linked to ADO',
                'Both Closed',
                'Both Open',
                'Status Mismatches',
                'Severity Mismatches',
                'Assignee Mismatches'
            ],
            'Count': [
                total_issues,
                linked_issues,
                unlinked_issues,
                both_closed,
                both_open,
                status_mismatches,
                severity_mismatches,
                assignee_mismatches
            ]
        }

        self.logger.info(f"Summary - Total: {total_issues}, Linked: {linked_issues}, Unlinked: {unlinked_issues}")

        return pd.DataFrame(summary_stats), {
            'total': total_issues,
            'linked': linked_issues,
            'unlinked': unlinked_issues,
            'status_mismatches': status_mismatches,
            'severity_mismatches': severity_mismatches,
            'assignee_mismatches': assignee_mismatches
        }

    def generate_excel_report(self, df, df_fuzzy_matches, summary_df):
        """Generate comprehensive Excel report"""
        self.logger.info("Generating Excel report...")

        # Prepare output directory and filename
        output_dir = Path(self.config['output']['report_directory'])
        output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = self.config['output']['report_filename'].format(timestamp=timestamp)
        output_file = output_dir / filename

        wb = Workbook()
        wb.remove(wb.active)

        # Sheet 1: Summary
        ws_summary = wb.create_sheet('Summary')
        ws_summary.append(['Jira-ADO Traceability Report'])
        ws_summary.append(['Generated on:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws_summary.append([])
        ws_summary.append(['Summary Statistics'])

        for r in dataframe_to_rows(summary_df, index=False, header=True):
            ws_summary.append(r)

        ws_summary['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws_summary['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        # Sheet 2: Full Traceability
        ws_full = wb.create_sheet('Full Traceability')
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_full.append(r)

        self._format_sheet_header(ws_full, '4472C4')
        self._auto_adjust_columns(ws_full)

        # Sheet 3: Mismatches
        df_mismatches = df[
            df['Status Comparison'].str.contains('[WARN]', na=False) |
            df['Severity Comparison'].str.contains('[WARN]', na=False) |
            df['Assignee Match'].str.contains('[WARN]', na=False)
        ]

        ws_mismatches = wb.create_sheet('Mismatches')
        for r in dataframe_to_rows(df_mismatches, index=False, header=True):
            ws_mismatches.append(r)

        self._format_sheet_header(ws_mismatches, 'C55A11')

        # Sheet 4: Matched Items
        df_matched = df[df['ADO ID'] != 'Not Linked'].copy()
        ws_matched = wb.create_sheet('Matched Items')
        for r in dataframe_to_rows(df_matched, index=False, header=True):
            ws_matched.append(r)

        self._format_sheet_header(ws_matched, '28A745')
        self._auto_adjust_columns(ws_matched)

        # Sheet 5: Potential Matches
        ws_fuzzy = wb.create_sheet('Potential Matches')
        if len(df_fuzzy_matches) > 0:
            for r in dataframe_to_rows(df_fuzzy_matches, index=False, header=True):
                ws_fuzzy.append(r)

            self._format_sheet_header(ws_fuzzy, 'FFA500')
            self._auto_adjust_columns(ws_fuzzy)
        else:
            ws_fuzzy.append(['No potential matches found'])

        # Sheet 6: Unlinked Issues
        df_unlinked = df[df['ADO ID'] == 'Not Linked']
        ws_unlinked = wb.create_sheet('Unlinked Issues')
        for r in dataframe_to_rows(df_unlinked[['Jira Key', 'Jira Summary', 'Jira Status', 'Jira Severity', 'Jira Assignee']], index=False, header=True):
            ws_unlinked.append(r)

        self._format_sheet_header(ws_unlinked, 'E74C3C')

        # Save workbook
        wb.save(output_file)

        self.logger.info(f"Report saved successfully: {output_file}")
        return str(output_file)

    @staticmethod
    def _format_sheet_header(worksheet, color):
        """Format worksheet header row"""
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

    @staticmethod
    def _auto_adjust_columns(worksheet):
        """Auto-adjust column widths"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def run(self):
        """Main execution method"""
        try:
            # Load Jira data
            jira_issues = self.load_jira_data()

            # Parse Jira issues
            df = self.parse_jira_issues(jira_issues)

            # Fetch ADO work items
            ado_work_items = self.fetch_ado_work_items(df)

            # Perform fuzzy matching
            df_fuzzy_matches = self.perform_fuzzy_matching(df)

            # Merge ADO data
            df = self.merge_ado_data(df, ado_work_items)

            # Perform comparison
            df = self.perform_comparison(df)

            # Generate summary
            summary_df, stats = self.generate_summary_stats(df)

            # Generate Excel report
            report_file = self.generate_excel_report(df, df_fuzzy_matches, summary_df)

            self.logger.info("="*80)
            self.logger.info("EXECUTION SUCCESSFUL")
            self.logger.info(f"Report: {report_file}")
            self.logger.info(f"Total Issues: {stats['total']}")
            self.logger.info(f"Linked: {stats['linked']}, Unlinked: {stats['unlinked']}")
            self.logger.info(f"Status Mismatches: {stats['status_mismatches']}")
            self.logger.info("="*80)

            return 0  # Success

        except Exception as e:
            self.logger.error("="*80)
            self.logger.error("EXECUTION FAILED")
            self.logger.error(f"Error: {str(e)}")
            self.logger.error(traceback.format_exc())
            self.logger.error("="*80)
            return 1  # Failure


def main():
    """Entry point for scheduled execution"""
    config_file = os.environ.get('JIRA_ADO_CONFIG',
                                 os.path.join(os.path.dirname(__file__), 'config.json'))

    try:
        robot = JiraAdoTraceabilityRobot(config_path=config_file)
        exit_code = robot.run()
        sys.exit(exit_code)

    except Exception as e:
        print(f"FATAL ERROR: {str(e)}", file=sys.stderr)
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
