import json
import pandas as pd
import requests
from requests.auth import HTTPBasicAuth
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# ===== CONFIGURATION =====
# Azure DevOps Configuration
ADO_SERVER = "http://tfsserver:8080/tfs"  # Your TFS server URL
ADO_COLLECTION = "YourCollection"          # e.g., "Iho"
ADO_PROJECT = "YourProject"                # e.g., "Guyana"
ADO_PAT = "your-ado-personal-access-token"  # Generate at: User Settings > Personal access tokens

# TFS/ADO API base URL
ADO_API_BASE = f"{ADO_SERVER}/{ADO_COLLECTION}/{ADO_PROJECT}/_apis"

# ===== LOAD JIRA DATA =====
print("Loading Jira data...")
with open(r"C:\Users\nkirue\AI-Playarea\jira_with_ado.json", "r", encoding="utf-8") as f:
    jira_data = json.load(f)

jira_issues = jira_data.get("issues", [])
print(f"Found {len(jira_issues)} Jira issues")

# ===== PARSE JIRA ISSUES =====
print("\nParsing Jira issues...")
parsed_data = []

for issue in jira_issues:
    fields = issue.get("fields", {})

    # Basic fields
    jira_key = issue.get("key", "")
    jira_summary = fields.get("summary", "")

    # Status
    status = fields.get("status", {})
    jira_status = status.get("name", "Unknown")
    jira_status_category = status.get("statusCategory", {}).get("name", "Unknown")

    # Priority
    priority = fields.get("priority", {})
    jira_priority = priority.get("name", "None") if priority else "None"

    # Severity (customfield_10042)
    severity = fields.get("customfield_10042", {})
    jira_severity = severity.get("value", "None") if severity else "None"

    # Assignee
    assignee = fields.get("assignee", {})
    jira_assignee = assignee.get("displayName", "Unassigned") if assignee else "Unassigned"

    # Dates
    created = fields.get("created", "")
    jira_created = pd.to_datetime(created, utc=True).tz_localize(None) if created else None

    resolutiondate = fields.get("resolutiondate", None)
    jira_resolved = pd.to_datetime(resolutiondate, utc=True).tz_localize(None) if resolutiondate else None

    # ADO fields
    ado_id = fields.get("customfield_10109", "")  # AzureDevOps ID
    ado_state_jira = fields.get("customfield_10110", "")  # AzureDevOps State (from Jira)

    parsed_data.append({
        "Jira Key": jira_key,
        "Jira Summary": jira_summary,
        "Jira Status": jira_status,
        "Jira Status Category": jira_status_category,
        "Jira Priority": jira_priority,
        "Jira Severity": jira_severity,
        "Jira Assignee": jira_assignee,
        "Jira Created": jira_created,
        "Jira Resolved": jira_resolved,
        "ADO ID": ado_id if ado_id else "Not Linked",
        "ADO State (Jira)": ado_state_jira if ado_state_jira else "N/A"
    })

df = pd.DataFrame(parsed_data)

# ===== FETCH ADO WORK ITEMS =====
print("\nFetching Azure DevOps work items...")

# Filter for issues that have ADO IDs
df_with_ado = df[df["ADO ID"] != "Not Linked"].copy()
ado_ids = df_with_ado["ADO ID"].unique()

print(f"Found {len(ado_ids)} unique ADO work items to fetch")

# Fetch ADO work items
ado_work_items = {}

if len(ado_ids) > 0:
    # Use Basic Auth with empty username and PAT as password for TFS
    auth = HTTPBasicAuth('', ADO_PAT)

    for ado_id in ado_ids:
        if not ado_id or ado_id == "Not Linked":
            continue

        try:
            # TFS 2018/Azure DevOps Server API endpoint
            url = f"{ADO_API_BASE}/wit/workitems/{ado_id}?api-version=5.0"

            print(f"Fetching ADO work item {ado_id}...")
            response = requests.get(url, auth=auth, timeout=10)

            if response.status_code == 200:
                work_item = response.json()
                fields = work_item.get("fields", {})

                ado_work_items[str(ado_id)] = {
                    "ADO Title": fields.get("System.Title", ""),
                    "ADO State": fields.get("System.State", ""),
                    "ADO Assigned To": fields.get("System.AssignedTo", {}).get("displayName", "Unassigned") if isinstance(fields.get("System.AssignedTo"), dict) else str(fields.get("System.AssignedTo", "Unassigned")),
                    "ADO Work Item Type": fields.get("System.WorkItemType", ""),
                    "ADO Priority": str(fields.get("Microsoft.VSTS.Common.Priority", "")),
                    "ADO Severity": str(fields.get("Microsoft.VSTS.Common.Severity", "")),
                    "ADO Created Date": fields.get("System.CreatedDate", ""),
                    "ADO Closed Date": fields.get("Microsoft.VSTS.Common.ClosedDate", ""),
                    "ADO Resolved Date": fields.get("Microsoft.VSTS.Common.ResolvedDate", ""),
                    "ADO Area Path": fields.get("System.AreaPath", ""),
                    "ADO Iteration Path": fields.get("System.IterationPath", "")
                }
                print(f"  [OK] Successfully fetched ADO-{ado_id}: {fields.get('System.State', 'Unknown')}")
            else:
                print(f"  [FAIL] Failed to fetch ADO-{ado_id}: HTTP {response.status_code}")
                ado_work_items[str(ado_id)] = {
                    "ADO Title": "ERROR: Could not fetch",
                    "ADO State": f"HTTP {response.status_code}",
                    "ADO Assigned To": "",
                    "ADO Work Item Type": "",
                    "ADO Priority": "",
                    "ADO Severity": "",
                    "ADO Created Date": "",
                    "ADO Closed Date": "",
                    "ADO Resolved Date": "",
                    "ADO Area Path": "",
                    "ADO Iteration Path": ""
                }
        except Exception as e:
            print(f"  [ERROR] Error fetching ADO-{ado_id}: {str(e)}")
            ado_work_items[str(ado_id)] = {
                "ADO Title": f"ERROR: {str(e)}",
                "ADO State": "Error",
                "ADO Assigned To": "",
                "ADO Work Item Type": "",
                "ADO Priority": "",
                "ADO Severity": "",
                "ADO Created Date": "",
                "ADO Closed Date": "",
                "ADO Resolved Date": "",
                "ADO Area Path": "",
                "ADO Iteration Path": ""
            }

# ===== MERGE ADO DATA WITH JIRA DATA =====
print("\nMerging Jira and ADO data...")

# Add ADO columns to DataFrame
ado_columns = ["ADO Title", "ADO State", "ADO Assigned To", "ADO Work Item Type",
               "ADO Priority", "ADO Severity", "ADO Created Date", "ADO Closed Date",
               "ADO Resolved Date", "ADO Area Path", "ADO Iteration Path"]

for col in ado_columns:
    df[col] = ""

# Fill in ADO data
for idx, row in df.iterrows():
    ado_id = str(row["ADO ID"])
    if ado_id in ado_work_items:
        for col in ado_columns:
            df.at[idx, col] = ado_work_items[ado_id].get(col, "")

# ===== COMPARISON LOGIC =====
print("\nPerforming comparison analysis...")

def compare_status(jira_status_cat, ado_state):
    """Compare if status is aligned between Jira and ADO"""
    if not ado_state or ado_state == "":
        return "No ADO Link"

    # Normalize states
    jira_done = jira_status_cat.lower() == "done"
    ado_closed = ado_state.lower() in ["closed", "resolved", "done", "removed"]

    if jira_done and ado_closed:
        return "[OK] Both Closed"
    elif not jira_done and not ado_closed:
        return "[OK] Both Open"
    elif jira_done and not ado_closed:
        return "[WARN] Jira Closed, ADO Open"
    else:
        return "[WARN] ADO Closed, Jira Open"

def compare_severity(jira_sev, ado_sev):
    """Compare severity between systems"""
    if not ado_sev or ado_sev == "":
        return "N/A"

    # Extract numbers from severity strings (e.g., "Sev-4" -> "4")
    jira_num = ''.join(filter(str.isdigit, str(jira_sev)))
    ado_num = str(ado_sev).strip()

    if jira_num == ado_num:
        return "[OK] Match"
    else:
        return f"[WARN] Mismatch (J:{jira_sev} vs A:{ado_sev})"

# Add comparison columns
df["Status Comparison"] = df.apply(lambda row: compare_status(row["Jira Status Category"], row["ADO State"]), axis=1)
df["Severity Comparison"] = df.apply(lambda row: compare_severity(row["Jira Severity"], row["ADO Severity"]), axis=1)
df["Assignee Match"] = df.apply(lambda row: "[OK] Match" if row["Jira Assignee"].lower() == row["ADO Assigned To"].lower() else "[WARN] Different", axis=1)

# ===== GENERATE SUMMARY STATISTICS =====
print("\nGenerating summary statistics...")

total_issues = len(df)
linked_issues = len(df[df["ADO ID"] != "Not Linked"])
unlinked_issues = total_issues - linked_issues

status_mismatches = len(df[df["Status Comparison"].str.contains("[WARN]", na=False)])
severity_mismatches = len(df[df["Severity Comparison"].str.contains("[WARN]", na=False)])
assignee_mismatches = len(df[df["Assignee Match"].str.contains("[WARN]", na=False)])

both_closed = len(df[df["Status Comparison"] == "[OK] Both Closed"])
both_open = len(df[df["Status Comparison"] == "[OK] Both Open"])

summary_stats = {
    "Metric": [
        "Total Jira Issues",
        "Linked to ADO",
        "Not Linked to ADO",
        "Both Closed",
        "Both Open",
        "Status Mismatches",
        "Severity Mismatches",
        "Assignee Mismatches"
    ],
    "Count": [
        total_issues,
        linked_issues,
        unlinked_issues,
        both_closed,
        both_open,
        status_mismatches,
        severity_mismatches,
        assignee_mismatches
    ]
}

summary_df = pd.DataFrame(summary_stats)

# ===== GENERATE EXCEL REPORT =====
print("\nGenerating Excel report...")

wb = Workbook()
wb.remove(wb.active)

# === SHEET 1: SUMMARY ===
ws_summary = wb.create_sheet("Summary")
ws_summary.append(["Jira-ADO Traceability Report"])
ws_summary.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
ws_summary.append([])
ws_summary.append(["Summary Statistics"])

for r in dataframe_to_rows(summary_df, index=False, header=True):
    ws_summary.append(r)

# Style header
ws_summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
ws_summary["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# === SHEET 2: FULL TRACEABILITY ===
ws_full = wb.create_sheet("Full Traceability")
for r in dataframe_to_rows(df, index=False, header=True):
    ws_full.append(r)

# Format header
for cell in ws_full[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Auto-adjust columns
for column in ws_full.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws_full.column_dimensions[column_letter].width = adjusted_width

# === SHEET 3: MISMATCHES ONLY ===
df_mismatches = df[
    df["Status Comparison"].str.contains("[WARN]", na=False) |
    df["Severity Comparison"].str.contains("[WARN]", na=False) |
    df["Assignee Match"].str.contains("[WARN]", na=False)
]

ws_mismatches = wb.create_sheet("Mismatches")
for r in dataframe_to_rows(df_mismatches, index=False, header=True):
    ws_mismatches.append(r)

# Format header
for cell in ws_mismatches[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# === SHEET 4: MATCHED ITEMS (ALL LINKED ITEMS) ===
# Matched = Has a valid ADO link (regardless of status/severity/assignee comparison)
df_matched = df[df["ADO ID"] != "Not Linked"].copy()

ws_matched = wb.create_sheet("Matched Items")
for r in dataframe_to_rows(df_matched, index=False, header=True):
    ws_matched.append(r)

# Format header
for cell in ws_matched[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Auto-adjust columns
for column in ws_matched.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws_matched.column_dimensions[column_letter].width = adjusted_width

# === SHEET 5: MATCHED SUMMARY ===
ws_matched_summary = wb.create_sheet("Matched Summary")
ws_matched_summary.append(["Matched (Linked) Items Summary Report"])
ws_matched_summary.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
ws_matched_summary.append([])

# Calculate matched statistics
total_matched = len(df_matched)
matched_closed = len(df_matched[df_matched["Jira Status Category"] == "Done"])
matched_open = len(df_matched[df_matched["Jira Status Category"] != "Done"])

# Comparison statistics for matched items
matched_status_ok = len(df_matched[df_matched["Status Comparison"].str.contains("[OK]", na=False)])
matched_status_warn = len(df_matched[df_matched["Status Comparison"].str.contains("[WARN]", na=False)])
matched_severity_ok = len(df_matched[df_matched["Severity Comparison"].str.contains("[OK]", na=False)])
matched_severity_warn = len(df_matched[df_matched["Severity Comparison"].str.contains("[WARN]", na=False)])
matched_assignee_ok = len(df_matched[df_matched["Assignee Match"].str.contains("[OK]", na=False)])
matched_assignee_warn = len(df_matched[df_matched["Assignee Match"].str.contains("[WARN]", na=False)])

# Perfect matches (all three criteria match)
perfect_matches = len(df_matched[
    (df_matched["Status Comparison"].str.contains("[OK]", na=False)) &
    (df_matched["Severity Comparison"].str.contains("[OK]", na=False)) &
    (df_matched["Assignee Match"].str.contains("[OK]", na=False))
])

# Status breakdown for matched items
matched_status_counts = df_matched["Jira Status"].value_counts()
matched_ado_state_counts = df_matched["ADO State"].value_counts()

# Severity breakdown for matched items
matched_severity_counts = df_matched["Jira Severity"].value_counts()

# Assignee breakdown for matched items
matched_assignee_counts = df_matched["Jira Assignee"].value_counts()

# Summary statistics
ws_matched_summary.append(["Overall Linked Items Statistics"])
ws_matched_summary.append(["Metric", "Count", "Percentage"])
ws_matched_summary.append(["Total Linked Items", total_matched, "100%"])
ws_matched_summary.append(["Jira: Closed/Done", matched_closed, f"{(matched_closed/total_matched*100):.1f}%" if total_matched > 0 else "0%"])
ws_matched_summary.append(["Jira: Open/In Progress", matched_open, f"{(matched_open/total_matched*100):.1f}%" if total_matched > 0 else "0%"])
ws_matched_summary.append([])

ws_matched_summary.append(["Comparison Quality (Linked Items)"])
ws_matched_summary.append(["Metric", "Count", "Percentage"])
ws_matched_summary.append(["Perfect Matches (All 3 Criteria)", perfect_matches, f"{(perfect_matches/total_matched*100):.1f}%" if total_matched > 0 else "0%"])
ws_matched_summary.append(["Status Matches", matched_status_ok, f"{(matched_status_ok/total_matched*100):.1f}%" if total_matched > 0 else "0%"])
ws_matched_summary.append(["Status Mismatches", matched_status_warn, f"{(matched_status_warn/total_matched*100):.1f}%" if total_matched > 0 else "0%"])
ws_matched_summary.append(["Severity Matches", matched_severity_ok, f"{(matched_severity_ok/total_matched*100):.1f}%" if total_matched > 0 else "0%"])
ws_matched_summary.append(["Severity Mismatches", matched_severity_warn, f"{(matched_severity_warn/total_matched*100):.1f}%" if total_matched > 0 else "0%"])
ws_matched_summary.append(["Assignee Matches", matched_assignee_ok, f"{(matched_assignee_ok/total_matched*100):.1f}%" if total_matched > 0 else "0%"])
ws_matched_summary.append(["Assignee Mismatches", matched_assignee_warn, f"{(matched_assignee_warn/total_matched*100):.1f}%" if total_matched > 0 else "0%"])
ws_matched_summary.append([])

# Jira Status breakdown
ws_matched_summary.append(["Jira Status Breakdown (Matched Items)"])
ws_matched_summary.append(["Status", "Count"])
for status, count in matched_status_counts.items():
    ws_matched_summary.append([status, count])
ws_matched_summary.append([])

# ADO State breakdown
ws_matched_summary.append(["ADO State Breakdown (Matched Items)"])
ws_matched_summary.append(["State", "Count"])
for state, count in matched_ado_state_counts.items():
    ws_matched_summary.append([state, count])
ws_matched_summary.append([])

# Severity breakdown
ws_matched_summary.append(["Severity Breakdown (Matched Items)"])
ws_matched_summary.append(["Severity", "Count"])
for severity, count in matched_severity_counts.items():
    ws_matched_summary.append([severity, count])
ws_matched_summary.append([])

# Top assignees
ws_matched_summary.append(["Top Assignees (Matched Items)"])
ws_matched_summary.append(["Assignee", "Count"])
for assignee, count in matched_assignee_counts.head(10).items():
    ws_matched_summary.append([assignee, count])

# Style the summary sheet
ws_matched_summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
ws_matched_summary["A1"].fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")

# Bold all section headers
for row in ws_matched_summary.iter_rows():
    if row[0].value and isinstance(row[0].value, str):
        if "breakdown" in row[0].value.lower() or "statistics" in row[0].value.lower() or "top" in row[0].value.lower():
            row[0].font = Font(bold=True, size=12)

# === SHEET 6: UNLINKED ISSUES ===
df_unlinked = df[df["ADO ID"] == "Not Linked"]

ws_unlinked = wb.create_sheet("Unlinked Issues")
for r in dataframe_to_rows(df_unlinked[["Jira Key", "Jira Summary", "Jira Status", "Jira Severity", "Jira Assignee"]], index=False, header=True):
    ws_unlinked.append(r)

# Format header
for cell in ws_unlinked[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Save workbook
output_file = r"C:\Users\nkirue\AI-Playarea\Jira_ADO_Traceability_Report.xlsx"
wb.save(output_file)

print(f"\n[SUCCESS] Report generated successfully: {output_file}")
print(f"\nSummary:")
print(f"  Total Issues: {total_issues}")
print(f"  Linked to ADO: {linked_issues}")
print(f"  Not Linked: {unlinked_issues}")
print(f"  Perfect Matches (Status+Severity+Assignee): {perfect_matches}")
print(f"  Status Mismatches (among linked): {status_mismatches}")
print(f"  Severity Mismatches (among linked): {severity_mismatches}")
print(f"  Assignee Mismatches (among linked): {assignee_mismatches}")
