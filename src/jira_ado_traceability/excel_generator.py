"""Excel report generation for Jira-ADO traceability."""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from jira_ado_traceability.models import FuzzyMatch


def create_workbook() -> Workbook:
    """Create and initialize workbook.

    Returns:
        New workbook instance
    """
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def format_header_row(sheet: any, row_num: int = 1, color: str = "4472C4") -> None:
    """Format header row with styling.

    Args:
        sheet: Worksheet to format
        row_num: Row number to format
        color: Hex color code for fill
    """
    for cell in sheet[row_num]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")


def auto_adjust_columns(sheet: any, max_width: int = 50) -> None:
    """Auto-adjust column widths based on content.

    Args:
        sheet: Worksheet to adjust
        max_width: Maximum column width
    """
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, max_width)
        sheet.column_dimensions[column_letter].width = adjusted_width


def add_summary_sheet(wb: Workbook, summary_df: pd.DataFrame) -> None:
    """Add summary statistics sheet.

    Args:
        wb: Workbook to add sheet to
        summary_df: DataFrame with summary statistics
    """
    ws = wb.create_sheet("Summary")
    ws.append(["Jira-ADO Traceability Report"])
    ws.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])
    ws.append(["Summary Statistics"])

    for row in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(row)

    # Style header
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


def add_full_traceability_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Add full traceability sheet with all data.

    Args:
        wb: Workbook to add sheet to
        df: Full traceability DataFrame
    """
    ws = wb.create_sheet("Full Traceability")
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    format_header_row(ws, row_num=1, color="4472C4")
    auto_adjust_columns(ws)


def add_mismatches_sheet(wb: Workbook, df_mismatches: pd.DataFrame) -> None:
    """Add mismatches sheet.

    Args:
        wb: Workbook to add sheet to
        df_mismatches: DataFrame with mismatched items
    """
    ws = wb.create_sheet("Mismatches")
    for row in dataframe_to_rows(df_mismatches, index=False, header=True):
        ws.append(row)

    format_header_row(ws, row_num=1, color="C55A11")


def add_matched_items_sheet(wb: Workbook, df_matched: pd.DataFrame) -> None:
    """Add matched items sheet.

    Args:
        wb: Workbook to add sheet to
        df_matched: DataFrame with matched items
    """
    ws = wb.create_sheet("Matched Items")
    for row in dataframe_to_rows(df_matched, index=False, header=True):
        ws.append(row)

    format_header_row(ws, row_num=1, color="28A745")
    auto_adjust_columns(ws)


def add_matched_summary_sheet(wb: Workbook, df_matched: pd.DataFrame) -> None:
    """Add matched summary sheet with detailed analytics.

    Args:
        wb: Workbook to add sheet to
        df_matched: DataFrame with matched items
    """
    ws = wb.create_sheet("Matched Summary")
    ws.append(["Matched (Linked) Items Summary Report"])
    ws.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])

    # Calculate statistics
    total_matched = len(df_matched)
    matched_closed = len(df_matched[df_matched["Jira Status Category"] == "Done"])
    matched_open = len(df_matched[df_matched["Jira Status Category"] != "Done"])

    # Comparison statistics
    matched_status_ok = len(df_matched[df_matched["Status Comparison"].str.contains("[OK]", na=False)])
    matched_status_warn = len(df_matched[df_matched["Status Comparison"].str.contains("[WARN]", na=False)])
    matched_severity_ok = len(df_matched[df_matched["Severity Comparison"].str.contains("[OK]", na=False)])
    matched_severity_warn = len(df_matched[df_matched["Severity Comparison"].str.contains("[WARN]", na=False)])
    matched_assignee_ok = len(df_matched[df_matched["Assignee Match"].str.contains("[OK]", na=False)])
    matched_assignee_warn = len(df_matched[df_matched["Assignee Match"].str.contains("[WARN]", na=False)])

    # Perfect matches
    perfect_matches = len(
        df_matched[
            (df_matched["Status Comparison"].str.contains("[OK]", na=False))
            & (df_matched["Severity Comparison"].str.contains("[OK]", na=False))
            & (df_matched["Assignee Match"].str.contains("[OK]", na=False))
        ]
    )

    # Add overall statistics
    ws.append(["Overall Linked Items Statistics"])
    ws.append(["Metric", "Count", "Percentage"])
    ws.append(["Total Linked Items", total_matched, "100%"])
    ws.append([
        "Jira: Closed/Done",
        matched_closed,
        f"{(matched_closed/total_matched*100):.1f}%" if total_matched > 0 else "0%",
    ])
    ws.append([
        "Jira: Open/In Progress",
        matched_open,
        f"{(matched_open/total_matched*100):.1f}%" if total_matched > 0 else "0%",
    ])
    ws.append([])

    # Add comparison quality
    ws.append(["Comparison Quality (Linked Items)"])
    ws.append(["Metric", "Count", "Percentage"])
    ws.append([
        "Perfect Matches (All 3 Criteria)",
        perfect_matches,
        f"{(perfect_matches/total_matched*100):.1f}%" if total_matched > 0 else "0%",
    ])
    ws.append([
        "Status Matches",
        matched_status_ok,
        f"{(matched_status_ok/total_matched*100):.1f}%" if total_matched > 0 else "0%",
    ])
    ws.append([
        "Status Mismatches",
        matched_status_warn,
        f"{(matched_status_warn/total_matched*100):.1f}%" if total_matched > 0 else "0%",
    ])
    ws.append([
        "Severity Matches",
        matched_severity_ok,
        f"{(matched_severity_ok/total_matched*100):.1f}%" if total_matched > 0 else "0%",
    ])
    ws.append([
        "Severity Mismatches",
        matched_severity_warn,
        f"{(matched_severity_warn/total_matched*100):.1f}%" if total_matched > 0 else "0%",
    ])
    ws.append([
        "Assignee Matches",
        matched_assignee_ok,
        f"{(matched_assignee_ok/total_matched*100):.1f}%" if total_matched > 0 else "0%",
    ])
    ws.append([
        "Assignee Mismatches",
        matched_assignee_warn,
        f"{(matched_assignee_warn/total_matched*100):.1f}%" if total_matched > 0 else "0%",
    ])
    ws.append([])

    # Add breakdowns
    _add_status_breakdown(ws, df_matched)
    _add_top_assignees(ws, df_matched)

    # Style the header
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")

    # Bold section headers
    for row in ws.iter_rows():
        if row[0].value and isinstance(row[0].value, str):
            if any(word in row[0].value.lower() for word in ["breakdown", "statistics", "top"]):
                row[0].font = Font(bold=True, size=12)


def _add_status_breakdown(ws: any, df_matched: pd.DataFrame) -> None:
    """Add status breakdown to sheet.

    Args:
        ws: Worksheet
        df_matched: Matched items DataFrame
    """
    ws.append(["Jira Status Breakdown (Matched Items)"])
    ws.append(["Status", "Count"])
    for status, count in df_matched["Jira Status"].value_counts().items():
        ws.append([status, count])
    ws.append([])

    ws.append(["ADO State Breakdown (Matched Items)"])
    ws.append(["State", "Count"])
    for state, count in df_matched["ADO State"].value_counts().items():
        ws.append([state, count])
    ws.append([])


def _add_top_assignees(ws: any, df_matched: pd.DataFrame) -> None:
    """Add top assignees to sheet.

    Args:
        ws: Worksheet
        df_matched: Matched items DataFrame
    """
    ws.append(["Top Assignees (Matched Items)"])
    ws.append(["Assignee", "Count"])
    for assignee, count in df_matched["Jira Assignee"].value_counts().head(10).items():
        ws.append([assignee, count])


def add_fuzzy_matches_sheet(wb: Workbook, fuzzy_matches: list[FuzzyMatch]) -> None:
    """Add potential fuzzy matches sheet.

    Args:
        wb: Workbook to add sheet to
        fuzzy_matches: List of fuzzy matches
    """
    ws = wb.create_sheet("Potential Matches")

    if len(fuzzy_matches) > 0:
        df_fuzzy = pd.DataFrame(fuzzy_matches)
        for row in dataframe_to_rows(df_fuzzy, index=False, header=True):
            ws.append(row)

        format_header_row(ws, row_num=1, color="FFA500")
        auto_adjust_columns(ws)

        # Add instructions at top
        ws.insert_rows(1, 3)
        ws["A1"] = "Potential Matches Based on Title Similarity"
        ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        ws["A2"] = (
            "Suggested matches using fuzzy text matching (Score >= 70%). "
            "Only Medium, High, and Very High confidence matches are shown."
        )
        ws["A2"].font = Font(italic=True)
        ws.merge_cells("A2:I2")
    else:
        ws.append(["No potential matches found"])
        ws.append(["All unlinked Jira items have no similar titles in Azure DevOps work items."])


def add_unlinked_issues_sheet(wb: Workbook, df_unlinked: pd.DataFrame) -> None:
    """Add unlinked issues sheet.

    Args:
        wb: Workbook to add sheet to
        df_unlinked: DataFrame with unlinked issues
    """
    ws = wb.create_sheet("Unlinked Issues")
    cols = ["Jira Key", "Jira Summary", "Jira Status", "Jira Severity", "Jira Assignee"]
    for row in dataframe_to_rows(df_unlinked[cols], index=False, header=True):
        ws.append(row)

    format_header_row(ws, row_num=1, color="E74C3C")


def generate_excel_report(
    output_file: str | Path,
    df: pd.DataFrame,
    summary_df: pd.DataFrame,
    fuzzy_matches: list[FuzzyMatch],
) -> None:
    """Generate complete Excel report.

    Args:
        output_file: Path to output Excel file
        df: Full traceability DataFrame
        summary_df: Summary statistics DataFrame
        fuzzy_matches: List of fuzzy matches
    """
    print("\nGenerating Excel report...")

    wb = create_workbook()

    # Create filtered DataFrames
    df_mismatches = df[
        df["Status Comparison"].str.contains("[WARN]", na=False)
        | df["Severity Comparison"].str.contains("[WARN]", na=False)
        | df["Assignee Match"].str.contains("[WARN]", na=False)
    ]
    df_matched = df[df["ADO ID"] != "Not Linked"].copy()
    df_unlinked = df[df["ADO ID"] == "Not Linked"]

    # Add sheets
    add_summary_sheet(wb, summary_df)
    add_full_traceability_sheet(wb, df)
    add_mismatches_sheet(wb, df_mismatches)
    add_matched_items_sheet(wb, df_matched)
    add_matched_summary_sheet(wb, df_matched)
    add_fuzzy_matches_sheet(wb, fuzzy_matches)
    add_unlinked_issues_sheet(wb, df_unlinked)

    # Save workbook
    wb.save(str(output_file))
    print(f"[SUCCESS] Report generated successfully: {output_file}")
